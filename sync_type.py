import pandas as pd

def sync_type_column(source_file, target_file):
    # 读取源文件和目标文件的预处理表格
    source_df = pd.read_excel(source_file, sheet_name='预处理')
    target_df = pd.read_excel(target_file, sheet_name='预处理')
    
    print(f"源文件记录数: {len(source_df)}")
    print(f"目标文件记录数: {len(target_df)}")
    
    # 创建进度项到类型的映射字典
    type_mapping = dict(zip(source_df['进度项'], source_df['类型']))
    
    # 更新目标文件中的类型列
    updated_count = 0
    for idx, row in target_df.iterrows():
        if row['进度项'] in type_mapping:
            target_df.at[idx, '类型'] = type_mapping[row['进度项']]
            updated_count += 1
    
    print(f"更新的记录数: {updated_count}")
    
    # 保存更新后的数据
    with pd.ExcelWriter(target_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        target_df.to_excel(writer, sheet_name='预处理', index=False)
    
    # 调整列宽
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    
    wb = load_workbook(target_file)
    ws = wb['预处理']
    
    # 自动调整每列的宽度
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(target_file)
    
    return updated_count

if __name__ == "__main__":
    source_file = "10-12 copy.xlsx"
    target_file = "last.xlsx"
    
    try:
        updated = sync_type_column(source_file, target_file)
        print(f"同步完成！已更新 {updated} 条记录的类型")
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}") 