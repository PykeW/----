import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def preprocess_table(input_file):
    # 读取Excel文件，不指定header
    df = pd.read_excel(input_file, header=None)
    
    # 检查第一行是否包含长字符串，如果是则删除
    if df.iloc[0].astype(str).str.len().max() > 20:  # 如果第一行有超过20个字符的单元格
        df = df.iloc[1:]  # 删除第一行
    
    # 将当前的第一行设为列名
    df.columns = df.iloc[0]
    df = df.iloc[1:].reset_index(drop=True)  # 删除作为列名的行并重置索引
    
    # 打印列名，用于调试
    print("列名:", df.columns.tolist())
    
    # 处理周报人列的空值，用前一个非空值填充
    df['周报人'] = df['周报人'].ffill()
    
    # 处理进度项列的内容，按换行符拆分
    rows = []
    for _, row in df.iterrows():
        if isinstance(row['订单项目.本周进度及问题反馈'], str):
            # 拆分内容并去除空行
            items = [item.strip() for item in row['订单项目.本周进度及问题反馈'].split('\n') if item.strip()]
            
            # 为每个拆分项创建新行
            for item in items:
                new_row = {
                    '周报人': row['周报人'],
                    '立项项目': row['订单项目.立项项目'],
                    '责任中心': row['订单项目.归属中心'],
                    '进度项': item,
                    '类型': '',  # 添加空的类型列
                    '记录ID': row['订单项目.记录ID(不可修改)']  # 放在最后一列
                }
                rows.append(new_row)
    
    # 创建新的DataFrame，并指定列的顺序
    result_df = pd.DataFrame(rows, columns=['周报人', '立项项目', '责任中心', '进度项', '类型', '记录ID'])
    
    # 使用ExcelWriter保存到原文件的新表格中
    with pd.ExcelWriter(input_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='预处理', index=False)
    
    # 打开工作簿并自动调整列宽
    wb = load_workbook(input_file)
    ws = wb['预处理']
    
    # 自动调整每列的宽度
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # 设置列宽（稍微加宽一点以确保完全显示）
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存更改
    wb.save(input_file)
    
    return result_df

if __name__ == "__main__":
    input_file = "2024三季度周报.xlsx"
    
    try:
        result = preprocess_table(input_file)
        print(f"处理完成！结果已保存到 {input_file} 的'预处理'表格中")
        print(f"共处理 {len(result)} 条记录")
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}") 