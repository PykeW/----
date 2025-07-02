import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate_plan_list(input_file):
    # 读取预处理表格
    df = pd.read_excel(input_file, sheet_name='预处理')
    
    # 创建结果DataFrame，设置所需的列
    result_columns = [
        '序号', '项目类型', '项目名称', '项目内容', 
        '责任部门', '项目经理', '计划开始时间', 
        '计划结束时间', '备注'
    ]
    result_df = pd.DataFrame(columns=result_columns)
    
    # 按类型和责任中心分组处理数据
    grouped = df.groupby(['类型', '责任中心'])
    
    current_index = 1  # 用于生成序号
    
    for (type_name, center), group in grouped:
        # 按立项项目分组，合并相同项目的进度项
        project_groups = group.groupby('立项项目')
        
        for project_name, project_group in project_groups:
            # 合并所有进度项作为项目内容
            project_content = '\n'.join(project_group['进度项'].tolist())
            
            # 获取项目经理（周报人）
            project_manager = project_group['周报人'].iloc[0]
            
            # 创建新行
            new_row = {
                '序号': current_index,
                '项目类型': type_name,
                '项目名称': project_name,
                '项目内容': project_content,
                '责任部门': center,
                '项目经理': project_manager,
                '计划开始时间': '',  # 留空，后续手动填写
                '计划结束时间': '',  # 留空，后续手动填写
                '备注': ''
            }
            
            result_df = pd.concat([result_df, pd.DataFrame([new_row])], ignore_index=True)
            current_index += 1
    
    # 保存结果到原文件的新表格中
    with pd.ExcelWriter(input_file, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
        result_df.to_excel(writer, sheet_name='计划清单', index=False)
    
    # 调整列宽
    wb = load_workbook(input_file)
    ws = wb['计划清单']
    
    # 自动调整每列的宽度
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # 限制最大宽度为50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(input_file)
    
    return len(result_df)

if __name__ == "__main__":
    input_file = "last.xlsx"
    
    try:
        record_count = generate_plan_list(input_file)
        print(f"处理完成！已生成 {record_count} 条计划记录")
        print(f"结果已保存到 {input_file} 的'计划清单'表格中")
    except Exception as e:
        print(f"处理过程中出现错误：{str(e)}")