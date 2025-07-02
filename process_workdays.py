import pandas as pd
import calendar
from datetime import datetime
import logging
import re
from openpyxl.styles import Border, Side, Alignment, PatternFill

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def get_quarter(month):
    return (month - 1) // 3 + 1

def extract_month_week(title):
    """从数据标题中提取月份和星期信息"""
    try:
        # 处理空值或非字符串值
        if pd.isna(title) or not isinstance(title, str):
            return None, None
            
        # 假设格式为: "陆杰    12    52" 这样的格式
        parts = re.split(r'\s+', title.strip())
        if len(parts) >= 3:
            try:
                return int(parts[1]), int(parts[2])
            except (ValueError, IndexError):
                return None, None
        return None, None
    except Exception as e:
        logging.warning(f"提取月份和星期时出错: {str(e)}, 标题内容: {title}")
        return None, None

def process_workdays_data(input_file):
    try:
        # 读取Excel文件，跳过第一行，使用第二行作为表头
        logging.info(f"开始读取文件: {input_file}")
        df = pd.read_excel(input_file, header=1)
        
        # 打印列名，用于调试
        logging.info("Excel文件包含以下列:")
        for col in df.columns:
            logging.info(f"- {col}")
        
        # 检查必需的列是否存在
        required_columns = ['周报人', '订单项目.立项项目', '订单项目.归属中心', 
                          '订单项目.本周投入天数（最低半天）', '数据标题(不可修改)']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"缺少必需的列: {missing_columns}")
        
        # 从数据标题中提取月份和星期信息
        logging.info("从数据标题中提取月份和星期信息...")
        month_week_data = df['数据标题(不可修改)'].apply(extract_month_week).tolist()
        df[['月份', '星期']] = pd.DataFrame(month_week_data, index=df.index)
        
        # 检查是否成功提取了月份信息
        if df['月份'].isna().all():
            raise ValueError("无法从数据标题中提取月份信息")
            
        # 使用2024年和月份数据生成日期
        logging.info("处理日期数据...")
        df['年份'] = 2024
        df['日期'] = pd.to_datetime(df.apply(
            lambda row: f"2024-{int(row['月份']):02d}-01" if pd.notna(row['月份']) else None, 
            axis=1
        ))
        
        # 移除没有有效日期的行
        df = df.dropna(subset=['日期'])
        if df.empty:
            raise ValueError("处理后没有有效数据")
        
        logging.info("开始处理数据...")
        # 添加季度列（月份已存在）
        df['季度'] = df['月份'].apply(get_quarter)
        
        days_column = '订单项目.本周投入天数（最低半天）'
        
        # 1. 按月度统计
        logging.info("开始按月度统计...")
        monthly_data = []
        
        # 按项目分组
        project_groups = df.groupby(['订单项目.归属中心', '订单项目.立项项目'])
        
        for (center, project), group in project_groups:
            months = sorted(group['月份'].unique())
            
            for month in months:
                month_data = group[group['月份'] == month]
                total_days = month_data[days_column].sum()
                
                if total_days > 0:
                    # 获取该月份的人员工时并排序
                    person_stats = month_data.groupby('周报人')[days_column].sum()
                    person_stats = person_stats[person_stats > 0].sort_values(ascending=False)
                    
                    # 获取第一个人员信息
                    first_person = person_stats.index[0] if not person_stats.empty else ''
                    first_days = person_stats.iloc[0] if not person_stats.empty else None
                    
                    # 添加项目行（包含第一个人员）
                    row_data = {
                        '订单项目.归属中心': center,
                        '订单项目.立项项目': project,
                        '月份': month,
                        '总人天': total_days,
                        '人员': first_person,
                        '人天': first_days
                    }
                    monthly_data.append(row_data)
                    
                    # 添加其余人员的工时
                    for person, days in list(person_stats.items())[1:]:
                        row_data = {
                            '订单项目.归属中心': center,
                            '订单项目.立项项目': project,
                            '月份': month,
                            '总人天': '',
                            '人员': person,
                            '人天': days
                        }
                        monthly_data.append(row_data)
        
        monthly_final = pd.DataFrame(monthly_data)
        monthly_final = monthly_final.sort_values(['订单项目.归属中心', '订单项目.立项项目', '月份'])
        
        # 2. 按季度统计（类似处理）
        logging.info("开始按季度统计...")
        quarterly_data = []
        
        for (center, project), group in project_groups:
            quarters = sorted(group['季度'].unique())
            
            for quarter in quarters:
                quarter_data = group[group['季度'] == quarter]
                total_days = quarter_data[days_column].sum()
                
                if total_days > 0:
                    person_stats = quarter_data.groupby('周报人')[days_column].sum()
                    person_stats = person_stats[person_stats > 0].sort_values(ascending=False)
                    
                    # 获取第一个人员信息
                    first_person = person_stats.index[0] if not person_stats.empty else ''
                    first_days = person_stats.iloc[0] if not person_stats.empty else None
                    
                    quarterly_data.append({
                        '订单项目.归属中心': center,
                        '订单项目.立项项目': project,
                        '季度': quarter,
                        '总人天': total_days,
                        '人员': first_person,
                        '人天': first_days
                    })
                    
                    # 添加其余人员
                    for person, days in list(person_stats.items())[1:]:
                        quarterly_data.append({
                            '订单项目.归属中心': center,
                            '订单项目.立项项目': project,
                            '季度': quarter,
                            '总人天': '',
                            '人员': person,
                            '人天': days
                        })
        
        quarterly_final = pd.DataFrame(quarterly_data)
        quarterly_final = quarterly_final.sort_values(['订单项目.归属中心', '订单项目.立项项目', '季度'])
        
        # 3. 按年度统计（类似处理）
        logging.info("开始按年度统计...")
        yearly_data = []
        
        for (center, project), group in project_groups:
            total_days = group[days_column].sum()
            
            if total_days > 0:
                person_stats = group.groupby('周报人')[days_column].sum()
                person_stats = person_stats[person_stats > 0].sort_values(ascending=False)
                
                yearly_data.append({
                    '订单项目.归属中心': center,
                    '订单项目.立项项目': project,
                    '总人天': total_days,
                    '人员': '',
                    '人天': ''
                })
                
                for person, days in person_stats.items():
                    yearly_data.append({
                        '订单项目.归属中心': center,
                        '订单项目.立项项目': project,
                        '总人天': '',
                        '人员': person,
                        '人天': days
                    })
        
        yearly_final = pd.DataFrame(yearly_data)
        yearly_final = yearly_final.sort_values(['订单项目.归属中心', '订单项目.立项项目'])

        logging.info(f"开始写入结果到文件...")
        with pd.ExcelWriter(input_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            write_with_merge(monthly_final, writer, '月度统计', ['订单项目.归属中心', '订单项目.立项项目'])
            write_with_merge(quarterly_final, writer, '季度统计', ['订单项目.归属中心', '订单项目.立项项目'])
            write_with_merge(yearly_final, writer, '年度统计', ['订单项目.归属中心', '订单项目.立项项目'])
        
        logging.info("处理完成!")
        
    except Exception as e:
        logging.error(f"处理过程中发生错误: {str(e)}", exc_info=True)
        raise

def write_with_merge(df, writer, sheet_name, merge_columns):
    """写入数据并合并相同内容的单元格"""
    # 写入数据
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    worksheet = writer.sheets[sheet_name]
    
    # 设置表头和单元格格式
    header_fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    # 设置所有单元格的基本格式
    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # 设置表头样式
    for cell in worksheet[1]:
        cell.fill = header_fill
    
    # 合并相同内容的单元格
    for col in merge_columns:
        if col not in df.columns:
            continue
            
        col_idx = df.columns.get_loc(col)
        current_value = None
        merge_start = 2
        
        for row in range(2, len(df) + 2):
            cell_value = worksheet.cell(row=row, column=col_idx + 1).value
            if cell_value != current_value:
                if current_value is not None and merge_start < row - 1:
                    worksheet.merge_cells(
                        start_row=merge_start,
                        start_column=col_idx + 1,
                        end_row=row - 1,
                        end_column=col_idx + 1
                    )
                merge_start = row
                current_value = cell_value
        
        # 处理最后一组
        if current_value is not None and merge_start < len(df) + 1:
            worksheet.merge_cells(
                start_row=merge_start,
                start_column=col_idx + 1,
                end_row=len(df) + 1,
                end_column=col_idx + 1
            )
    
    # 调整列宽
    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

if __name__ == '__main__':
    input_file = '周报汇总.xlsx'
    process_workdays_data(input_file) 